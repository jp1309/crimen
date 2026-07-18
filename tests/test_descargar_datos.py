from __future__ import annotations

import tempfile
import unittest
import shutil
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from scripts.descargar_datos import (
    SourceUpdateError,
    clasificar_recursos,
    sha256_file,
    sincronizar,
    validar_excel,
)


def resource(name: str, resource_id: str) -> dict[str, object]:
    return {
        "id": resource_id,
        "name": name,
        "format": "XLSX",
        "state": "active",
        "url": f"https://example.invalid/{name}",
    }


class ResourceClassificationTests(unittest.TestCase):
    def test_selects_historical_and_current_accumulated_files(self) -> None:
        resources = [
            resource("mdi_homicidios_intencionales_dd_2025.xlsx", "dictionary"),
            resource("mdi_homicidios_intencionales_pm_2014-2025.xlsx", "historical"),
            resource("mdi_homicidios_intencionales_pm_2026_enero_junio.xlsx", "current"),
        ]

        selected = clasificar_recursos(resources)

        self.assertEqual(selected["historico"]["id"], "historical")
        self.assertEqual(selected["actual"]["id"], "current")

    def test_fails_closed_when_a_primary_source_is_missing(self) -> None:
        with self.assertRaises(SourceUpdateError):
            clasificar_recursos(
                [resource("mdi_homicidios_intencionales_pm_2026_enero_junio.xlsx", "current")]
            )


class ExcelValidationTests(unittest.TestCase):
    def test_validates_a_workbook_with_official_headers(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "source.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Datos"
            worksheet.append(["PROVINCIA", "CANTON", "FECHA_INFRACCION"])
            for index in range(6_000):
                worksheet.append(["PICHINCHA", "QUITO", f"2026-01-{index % 28 + 1:02d}"])
            workbook.save(path)

            validation = validar_excel(path)

            self.assertEqual(validation["sheet"], "Datos")
            self.assertEqual(validation["header_row"], 1)
            self.assertEqual(len(sha256_file(path)), 64)


class AccumulatedReplacementTests(unittest.TestCase):
    @staticmethod
    def _workbook(path: Path, rows: list[tuple[str, str, str]]) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["PROVINCIA", "CANTON", "FECHA_INFRACCION"])
        for row in rows:
            worksheet.append(row)
        workbook.save(path)

    def test_current_year_is_replaced_not_appended_and_second_run_is_noop(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            project = Path(directory)
            raw_dir = project / "data" / "raw"
            raw_dir.mkdir(parents=True)
            manifest = project / "data" / "source_manifest.json"
            old_current = raw_dir / "mdi_homicidios_intencionales_pm_actual.xlsx"
            historical = project / "remote-historical.xlsx"
            new_current = project / "remote-current.xlsx"

            self._workbook(
                old_current,
                [("PICHINCHA", "QUITO", "2026-01-01"), ("GUAYAS", "GUAYAQUIL", "2026-02-01")],
            )
            self._workbook(historical, [("PICHINCHA", "QUITO", "2025-01-01")])
            self._workbook(
                new_current,
                [
                    ("PICHINCHA", "QUITO", "2026-01-02"),
                    ("GUAYAS", "GUAYAQUIL", "2026-02-01"),
                    ("AZUAY", "CUENCA", "2026-03-01"),
                ],
            )
            metadata = {
                "id": "dataset",
                "name": "homicidios-intencionales",
                "metadata_modified": "2026-04-15T00:00:00",
                "update_frequency": ["Mensual"],
                "resources": [
                    resource("mdi_homicidios_intencionales_pm_2014-2025.xlsx", "historical"),
                    resource("mdi_homicidios_intencionales_pm_2026_enero_marzo.xlsx", "current"),
                ],
            }
            downloads = {
                "https://example.invalid/mdi_homicidios_intencionales_pm_2014-2025.xlsx": historical,
                "https://example.invalid/mdi_homicidios_intencionales_pm_2026_enero_marzo.xlsx": new_current,
            }

            def fake_download(url: str, destination: Path) -> None:
                shutil.copyfile(downloads[url], destination)

            with (
                patch("scripts.descargar_datos.obtener_metadata", return_value=metadata),
                patch("scripts.descargar_datos.descargar_archivo", side_effect=fake_download),
                patch("scripts.descargar_datos.MIN_XLSX_SIZE", 100),
            ):
                changed, _ = sincronizar(raw_dir=raw_dir, manifest_path=manifest)
                first_manifest_hash = sha256_file(manifest)
                changed_again, _ = sincronizar(raw_dir=raw_dir, manifest_path=manifest)

            workbook = load_workbook(old_current, read_only=True, data_only=True)
            try:
                rows = list(workbook.active.iter_rows(min_row=2, values_only=True))
            finally:
                workbook.close()

            self.assertTrue(changed)
            self.assertFalse(changed_again)
            self.assertEqual(rows, [
                ("PICHINCHA", "QUITO", "2026-01-02"),
                ("GUAYAS", "GUAYAQUIL", "2026-02-01"),
                ("AZUAY", "CUENCA", "2026-03-01"),
            ])
            self.assertEqual(first_manifest_hash, sha256_file(manifest))


if __name__ == "__main__":
    unittest.main()
