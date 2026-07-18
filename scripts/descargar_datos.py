"""Sincroniza las fuentes oficiales de homicidios desde Datos Abiertos Ecuador.

El Ministerio reemplaza mensualmente el archivo acumulado del anio en curso y
puede corregir cualquiera de los meses ya publicados. Por eso este modulo no
anexa el ultimo mes: reemplaza el Excel anual completo despues de validarlo.

Uso:
    python -m scripts.descargar_datos
    python -m scripts.descargar_datos --dry-run
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import subprocess
import sys
import tempfile
import urllib.error
import urllib.request
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from scripts.configuracion import (
    CANONICAL_SOURCES,
    CKAN_DATASET_API,
    CKAN_DATASET_ID,
    DATA_RAW,
    SOURCE_MANIFEST,
    STABLE_RESOURCES,
)


USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
)
SOURCE_PAGE = "https://www.datosabiertos.gob.ec/dataset/homicidios-intencionales"
MIN_XLSX_SIZE = 50_000
YEAR_RANGE_RE = re.compile(r"20\d{2}[-_]20\d{2}")


class SourceUpdateError(RuntimeError):
    """Error controlado al consultar o validar una fuente oficial."""


def _request(url: str) -> urllib.request.Request:
    return urllib.request.Request(
        url,
        headers={
            "Accept": "application/json, application/octet-stream",
            "Accept-Language": "es-EC,es;q=0.9,en;q=0.8",
            "Referer": SOURCE_PAGE,
            "User-Agent": USER_AGENT,
        },
    )


def _curl(url: str, output: Path | None = None, timeout: int = 180) -> bytes:
    """Cliente alternativo para portales que bloquean la firma TLS de urllib."""

    command = [
        "curl",
        "--fail",
        "--location",
        "--retry",
        "2",
        "--retry-all-errors",
        "--silent",
        "--show-error",
        "--connect-timeout",
        "30",
        "--max-time",
        str(timeout),
        "--user-agent",
        USER_AGENT,
        "--referer",
        SOURCE_PAGE,
        "--header",
        "Accept-Language: es-EC,es;q=0.9,en;q=0.8",
    ]
    if output is not None:
        command.extend(["--output", str(output)])
    command.append(url)
    try:
        result = subprocess.run(
            command,
            check=True,
            capture_output=True,
            timeout=timeout + 15,
        )
    except (FileNotFoundError, subprocess.CalledProcessError, subprocess.TimeoutExpired) as exc:
        detail = ""
        if isinstance(exc, subprocess.CalledProcessError):
            detail = exc.stderr.decode("utf-8", errors="replace").strip()
        raise SourceUpdateError(f"curl no pudo descargar {url}: {detail or exc}") from exc
    return result.stdout


def obtener_metadata(api_url: str = CKAN_DATASET_API) -> dict[str, Any]:
    """Consulta CKAN y devuelve el objeto del conjunto de datos."""

    try:
        with urllib.request.urlopen(_request(api_url), timeout=60) as response:
            payload = json.load(response)
    except urllib.error.HTTPError as exc:
        if exc.code != 403:
            raise SourceUpdateError(f"No se pudo consultar la API oficial: {exc}") from exc
        print("La API rechazo urllib; reintentando con curl...")
        try:
            payload = json.loads(_curl(api_url, timeout=60))
        except (SourceUpdateError, json.JSONDecodeError) as curl_exc:
            raise SourceUpdateError(f"No se pudo consultar la API oficial: {curl_exc}") from curl_exc
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
        raise SourceUpdateError(f"No se pudo consultar la API oficial: {exc}") from exc

    if not payload.get("success") or not isinstance(payload.get("result"), dict):
        raise SourceUpdateError("La API oficial devolvio una respuesta no valida")
    return payload["result"]


def clasificar_recursos(resources: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """Identifica los dos XLSX primarios: historico y anio en curso."""

    selected: dict[str, dict[str, Any]] = {}
    for resource in resources:
        name = str(resource.get("name") or "")
        normalized = name.lower()
        if resource.get("state") != "active" or str(resource.get("format")).upper() != "XLSX":
            continue
        if "_pm_" not in normalized or "_dd_" in normalized:
            continue

        key = "historico" if YEAR_RANGE_RE.search(normalized) else "actual"
        if key in selected:
            raise SourceUpdateError(
                f"La API publico mas de un recurso primario para '{key}': "
                f"{selected[key].get('name')} y {name}"
            )
        if not resource.get("url") or not resource.get("id"):
            raise SourceUpdateError(f"El recurso {name!r} no tiene URL o identificador")
        selected[key] = resource

    missing = sorted(set(CANONICAL_SOURCES) - set(selected))
    if missing:
        raise SourceUpdateError(
            "No se encontraron todas las fuentes primarias en la API: " + ", ".join(missing)
        )
    return selected


def _load_manifest(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def recursos_estables(manifest_path: Path = SOURCE_MANIFEST) -> dict[str, dict[str, Any]]:
    """Construye recursos descargables aun cuando CKAN bloquee su API JSON.

    CKAN mantiene estables los identificadores y la ruta ``/download`` aunque
    cambie el nombre mensual del archivo. El manifiesto conserva el ultimo
    nombre descriptivo conocido.
    """

    previous = _load_manifest(manifest_path).get("resources") or {}
    resources: dict[str, dict[str, Any]] = {}
    for key, stable in STABLE_RESOURCES.items():
        previous_resource = previous.get(key) or {}
        resources[key] = {
            "id": stable["id"],
            "name": previous_resource.get("official_name") or CANONICAL_SOURCES[key].name,
            "format": "XLSX",
            "state": "active",
            "url": stable["url"],
            "last_modified": previous_resource.get("last_modified"),
        }
    return resources


def descargar_archivo(url: str, destination: Path) -> None:
    """Descarga una URL por bloques, sin cargar todo el Excel en memoria."""

    try:
        with urllib.request.urlopen(_request(url), timeout=180) as response:
            content_type = (response.headers.get("Content-Type") or "").lower()
            if response.status != 200:
                raise SourceUpdateError(f"Descarga HTTP {response.status}: {url}")
            if "spreadsheet" not in content_type and "octet-stream" not in content_type:
                raise SourceUpdateError(
                    f"Tipo de contenido inesperado ({content_type or 'sin tipo'}): {url}"
                )
            with destination.open("wb") as output:
                while chunk := response.read(1024 * 1024):
                    output.write(chunk)
    except urllib.error.HTTPError as exc:
        if exc.code != 403:
            raise SourceUpdateError(f"No se pudo descargar {url}: {exc}") from exc
        print("El recurso rechazo urllib; reintentando con curl...")
        _curl(url, output=destination)
    except (urllib.error.URLError, TimeoutError, OSError) as exc:
        raise SourceUpdateError(f"No se pudo descargar {url}: {exc}") from exc


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        while chunk := source.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def validar_excel(path: Path) -> dict[str, Any]:
    """Comprueba estructura XLSX y localiza una tabla oficial con PROVINCIA."""

    size = path.stat().st_size
    if size < MIN_XLSX_SIZE:
        raise SourceUpdateError(f"El archivo {path.name} es demasiado pequeno ({size} bytes)")
    if not zipfile.is_zipfile(path):
        raise SourceUpdateError(f"El archivo {path.name} no es un XLSX valido")

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                for row_number, row in enumerate(
                    worksheet.iter_rows(min_row=1, max_row=50, values_only=True), start=1
                ):
                    values = " ".join(str(value).upper() for value in row if value is not None)
                    if "PROVINCIA" in values and any(
                        token in values for token in ("FECHA", "ZONA", "CANTON")
                    ):
                        return {
                            "size": size,
                            "sheet": worksheet.title,
                            "header_row": row_number,
                        }
        finally:
            workbook.close()
    except Exception as exc:  # openpyxl expone varios tipos de error de archivo
        raise SourceUpdateError(f"No se pudo validar {path.name}: {exc}") from exc

    raise SourceUpdateError(f"No se encontro una tabla con PROVINCIA en {path.name}")


def _write_json_atomic(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_suffix(path.suffix + ".tmp")
    temp_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    os.replace(temp_path, path)


def _write_github_output(path: str | None, changed: bool, current_name: str) -> None:
    if not path:
        return
    with open(path, "a", encoding="utf-8") as output:
        output.write(f"changed={'true' if changed else 'false'}\n")
        output.write(f"source_name={current_name}\n")


def sincronizar(
    api_url: str = CKAN_DATASET_API,
    raw_dir: Path = DATA_RAW,
    manifest_path: Path = SOURCE_MANIFEST,
    dry_run: bool = False,
) -> tuple[bool, dict[str, Any]]:
    """Descarga, valida y reemplaza atomicamente las fuentes que cambiaron."""

    try:
        metadata = obtener_metadata(api_url)
        resources = clasificar_recursos(metadata.get("resources") or [])
    except SourceUpdateError as exc:
        # Datos Abiertos puede responder 403 a la API JSON desde rangos de IP
        # de GitHub Actions. Las URLs estables de los recursos siguen siendo la
        # fuente oficial y el SHA-256 determina si hubo una publicacion nueva.
        print(f"Advertencia: {exc}")
        print("Se usaran las URLs estables de los recursos oficiales.")
        resources = recursos_estables(manifest_path)
        metadata = {
            "id": CKAN_DATASET_ID,
            "name": "homicidios-intencionales",
            "metadata_modified": None,
            "update_frequency": ["Mensual"],
            "resources": list(resources.values()),
        }
    raw_dir.mkdir(parents=True, exist_ok=True)

    temporary: dict[str, Path] = {}
    prepared: dict[str, dict[str, Any]] = {}
    try:
        # Primero se descargan y validan ambas fuentes. Ningun archivo local se
        # reemplaza si una de ellas falla.
        for key, resource in resources.items():
            handle, temp_name = tempfile.mkstemp(
                prefix=f".download-{key}-", suffix=".xlsx", dir=raw_dir
            )
            os.close(handle)
            temp_path = Path(temp_name)
            temporary[key] = temp_path
            print(f"Descargando {resource['name']}...")
            # Se descarga por ID estable. La URL continua funcionando cuando
            # CKAN reemplaza el nombre del Excel en una nueva publicacion.
            download_url = STABLE_RESOURCES[key]["url"]
            descargar_archivo(download_url, temp_path)
            validation = validar_excel(temp_path)
            digest = sha256_file(temp_path)
            target = raw_dir / CANONICAL_SOURCES[key].name
            prepared[key] = {
                "resource": resource,
                "download_url": download_url,
                "temp_path": temp_path,
                "target": target,
                "sha256": digest,
                "validation": validation,
                "changed": not target.exists() or sha256_file(target) != digest,
            }

        changed = any(item["changed"] for item in prepared.values()) or not manifest_path.exists()
        if dry_run:
            print("Dry-run: no se reemplazaron archivos locales.")
            return changed, metadata

        if not changed:
            print("Las fuentes locales ya coinciden con la publicacion oficial.")
            return False, metadata

        for key, item in prepared.items():
            if item["changed"]:
                os.replace(item["temp_path"], item["target"])
                temporary.pop(key, None)
                print(f"Actualizada fuente {key}: {item['target'].name}")

        manifest = {
            "schema_version": 1,
            "dataset": {
                "api_url": api_url,
                "id": metadata.get("id"),
                "name": metadata.get("name"),
                "metadata_modified": metadata.get("metadata_modified"),
                "update_frequency": metadata.get("update_frequency"),
            },
            "downloaded_at": datetime.now(timezone.utc).isoformat(),
            "resources": {},
        }
        for key, item in prepared.items():
            resource = item["resource"]
            target = item["target"]
            manifest["resources"][key] = {
                "resource_id": resource.get("id"),
                "official_name": resource.get("name"),
                "official_url": resource.get("url"),
                "download_url": item["download_url"],
                "last_modified": resource.get("last_modified"),
                "local_path": target.relative_to(manifest_path.parent.parent).as_posix(),
                "sha256": sha256_file(target),
                "size": target.stat().st_size,
                "sheet": item["validation"]["sheet"],
                "header_row": item["validation"]["header_row"],
            }
        _write_json_atomic(manifest_path, manifest)
        return True, metadata
    finally:
        for path in temporary.values():
            try:
                path.unlink(missing_ok=True)
            except OSError:
                pass


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--api-url", default=CKAN_DATASET_API)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--github-output", help="Ruta de GITHUB_OUTPUT")
    args = parser.parse_args(argv)

    try:
        changed, metadata = sincronizar(api_url=args.api_url, dry_run=args.dry_run)
        current_id = STABLE_RESOURCES["actual"]["id"]
        current_resource = next(
            (
                resource
                for resource in metadata.get("resources") or []
                if resource.get("id") == current_id
            ),
            {},
        )
        current_name = str(current_resource.get("name") or CANONICAL_SOURCES["actual"].name)
        _write_github_output(args.github_output, changed, current_name)
        print(f"Fuente anual oficial: {current_name}")
        print(f"Cambios detectados: {'si' if changed else 'no'}")
        return 0
    except SourceUpdateError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
